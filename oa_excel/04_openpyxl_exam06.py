import openpyxl as op
path = r"C:\workspace\oa_excel\output"
wb = op.load_workbook(path + "/test_ullim.xlsx")

#ws = wb.active # 활성화된 시트
ws = wb["울림"] # 특정시트
data1 = ws["C4"].value
data2 = ws.cell(row=4, column=3).value
print(data1, data2)
