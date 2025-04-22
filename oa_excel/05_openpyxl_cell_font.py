import openpyxl as op
# from openpyxl.styles import Font

path = r"C:\workspace\oa_excel\output"
wb = op.load_workbook(path + "/font_example.xlsx")
ws = wb.create_sheet("Font")

# 셀에 폰트 적용
cell = ws['A1']
cell.value = "폰트 테스트22222"
cell.font = op.styles.Font(name='맑은 고딕', size=24, bold=True, italic=True, color="00FF00")

path = r"C:\workspace\oa_excel\output"
wb.save(path + "/font_example.xlsx")