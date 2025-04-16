import openpyxl as op
path = r"C:\workspace\oa_excel\output"
wb = op.load_workbook(path + "/test_ullim.xlsx")

#ws = wb.active # 활성화된 시트
ws = wb["울림"] # 특정시트

data1 = ws["A2:C4"]
# for i in data1:
#     for j in i:
#         print(j.value)

data2 = ws.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3)
for i in data2:
    for j in i:
        print(j.value)

print("*"*30)

data2 = ws.iter_cols(min_row=2, max_row=4, min_col=1, max_col=3)
for i in data2:
    for j in i:
        print(j.value)

