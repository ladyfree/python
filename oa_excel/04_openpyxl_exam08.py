import openpyxl as op
path = r"C:\workspace\oa_excel\output"
wb = op.load_workbook(path + "/test_ullim.xlsx")

sh_list = wb.sheetnames

for sheet_name in sh_list:
    ws = wb[sheet_name]
    
    # 새 워크북 만들기
    new_wb = op.Workbook()
    new_ws = new_wb.active
    new_ws.title = sheet_name
    
    # 셀 데이터를 새 워크북으로 복사
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)
    
    # 파일로 저장 (시트 이름을 파일 이름으로 사용)
    filename = path +'/' + f"{sheet_name}.xlsx"
    new_wb.save(filename)
    print(f"'{filename}' 저장 완료")