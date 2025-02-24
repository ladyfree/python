import openpyxl

# 새 워크북 생성
wb = openpyxl.Workbook()

# 기본 시트 선택
ws = wb.active

# 셀에 값 입력
ws["A1"] = "Hello"
ws["B1"] = "World!"

# 엑셀 파일 저장
wb.save("new_file.xlsx")
