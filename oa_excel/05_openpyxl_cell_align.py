import openpyxl as op

path = r"C:\workspace\oa_excel\output"
wb = op.load_workbook(path + "/font_example.xlsx")
ws = wb.create_sheet("Alignment")

# 셀에 정렬 적용
cell = ws['A1']
cell.value = "가운데 정렬"
cell.alignment = op.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

ws.row_dimensions[1].height = 50
ws.column_dimensions['A'].width = 30

wb.save(path + "/font_example.xlsx")